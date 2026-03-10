import streamlit as st
import gspread
from google.oauth2.service_account import Credentials
from datetime import date, datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io
import uuid

# ══════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════
SCOPES = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
JORNADA_BASE = 8.0

EQUIPOS = {
    "EQUIPO 1": "AB887CX",
    "EQUIPO 2": "AH453YE",
}
ESTADOS_SERVICIO = ["PENDIENTE", "CONFIRMADO", "REALIZADO", "SUSPENDIDO", "REPROGRAMADO"]

st.set_page_config(page_title="Panel de Control", page_icon="⚙️", layout="wide")

# ══════════════════════════════════════════════════════════════════
# CSS
# ══════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

* { font-family: 'DM Sans', sans-serif; }
code, .mono { font-family: 'DM Mono', monospace; }

.stApp { background: #f1f3f6; }
[data-testid="stSidebar"] { background: #ffffff !important; border-right: 1px solid #e2e6ea; }

#MainMenu, footer, header { visibility: hidden; }
[data-testid="stToolbar"] { display: none; }

[data-testid="metric-container"] {
    background: #ffffff;
    border: 1px solid #e2e6ea;
    border-radius: 12px;
    padding: 16px;
}
[data-testid="metric-container"] label { color: #6b7280 !important; font-size: 12px !important; }
[data-testid="metric-container"] [data-testid="stMetricValue"] { color: #1e3a8a !important; font-size: 22px !important; font-weight: 700 !important; }

.stButton > button {
    background: #1e3a8a !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
    transition: all 0.2s !important;
}
.stButton > button:hover { background: #2745a7 !important; transform: translateY(-1px); }

.stTextInput > div > div > input,
.stNumberInput > div > div > input,
.stSelectbox > div > div,
.stDateInput > div > div > input,
.stTextArea textarea {
    background: #ffffff !important;
    border: 1px solid #d1d5db !important;
    color: #111827 !important;
    border-radius: 8px !important;
}

.stTabs [data-baseweb="tab-list"] { background: #e5e7eb; border-radius: 10px; padding: 4px; gap: 4px; }
.stTabs [data-baseweb="tab"] { color: #6b7280 !important; border-radius: 8px !important; font-weight: 500; }
.stTabs [aria-selected="true"] { background: #1e3a8a !important; color: white !important; }

[data-testid="stDataFrame"] { border-radius: 12px; overflow: hidden; }
.stDataFrame { background: #ffffff !important; }

h1, h2, h3 { color: #111827 !important; }
p, label { color: #374151 !important; }
hr { border-color: #e2e6ea !important; }

.stSuccess { background: #f0fdf4 !important; border: 1px solid #86efac !important; color: #166534 !important; border-radius: 8px !important; }
.stError { background: #fef2f2 !important; border: 1px solid #fca5a5 !important; color: #991b1b !important; border-radius: 8px !important; }
.stWarning { background: #fffbeb !important; border: 1px solid #fcd34d !important; color: #92400e !important; border-radius: 8px !important; }

/* Sidebar nav buttons */
[data-testid="stSidebar"] .stButton > button {
    background: transparent !important;
    color: #374151 !important;
    border: none !important;
    border-radius: 6px !important;
    font-weight: 500 !important;
    text-align: left !important;
    padding: 6px 10px !important;
    font-size: 13px !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background: #f3f4f6 !important;
    color: #111827 !important;
    transform: none !important;
}
[data-testid="stSidebar"] .stButton > button[kind="primary"] {
    background: #eff6ff !important;
    color: #1e3a8a !important;
    font-weight: 600 !important;
}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════
# GOOGLE SHEETS
# ══════════════════════════════════════════════════════════════════
@st.cache_resource
def get_sheet():
    creds_dict = st.secrets["gcp_service_account"]
    creds = Credentials.from_service_account_info(dict(creds_dict), scopes=SCOPES)
    client = gspread.authorize(creds)
    return client.open_by_key(st.secrets["sheet_id"])

def get_ws(name: str):
    sh = get_sheet()
    try:
        return sh.worksheet(name)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=name, rows=2000, cols=20)
        return ws


# ══════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════
def decimal_a_hhmm(val: float) -> str:
    neg = val < 0
    v = abs(val)
    h = int(v)
    m = int(round((v - h) * 60))
    if m == 60:
        h += 1; m = 0
    s = f"{h}h {str(m).zfill(2)}m"
    return f"-{s}" if neg else s

def parse_hora(h: str) -> datetime:
    return datetime.strptime(h.strip(), "%H:%M")

def calcular_horas(entrada: str, salida: str):
    h_ent = parse_hora(entrada)
    h_sal = parse_hora(salida)
    if h_sal <= h_ent:
        return None, None
    diff = h_sal - h_ent
    trabajado = diff.total_seconds() / 3600
    return trabajado, trabajado - JORNADA_BASE

MESES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
         "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]


# ══════════════════════════════════════════════════════════════════
# DATOS: USUARIOS
# ══════════════════════════════════════════════════════════════════
def cargar_usuarios() -> dict:
    ws = get_ws("usuarios")
    vals = ws.get_all_values()
    if not vals:
        default = {"Alejo": "1234", "Martin": "1234"}
        ws.append_rows([[u, p] for u, p in default.items()])
        return default
    return {r[0]: r[1] for r in vals if len(r) >= 2 and r[0]}

def validar_usuario(usuario: str, password: str) -> bool:
    return cargar_usuarios().get(usuario) == password


# ══════════════════════════════════════════════════════════════════
# DATOS: EMPLEADOS (Horarios)
# ══════════════════════════════════════════════════════════════════
COLS_EMP = ["id","nombre","activo","telefono","dni","zona","vehiculo","patente","observaciones"]

def cargar_empleados(solo_activos=False) -> list:
    ws = get_ws("empleados")
    records = ws.get_all_records()
    empleados = []
    for r in records:
        emp = {
            "id": str(r.get("id","")),
            "nombre": r.get("nombre",""),
            "activo": str(r.get("activo","1")) == "1",
            "telefono": r.get("telefono",""),
            "dni": r.get("dni",""),
            "zona": r.get("zona",""),
            "vehiculo": r.get("vehiculo",""),
            "patente": r.get("patente",""),
            "observaciones": r.get("observaciones",""),
        }
        empleados.append(emp)
    if solo_activos:
        empleados = [e for e in empleados if e["activo"]]
    return sorted(empleados, key=lambda x: x["nombre"])

def _init_empleados_ws():
    ws = get_ws("empleados")
    if not ws.get_all_values():
        ws.append_row(COLS_EMP)
        for nombre in ["Maxi","Sergio","Hugo","Lautaro"]:
            ws.append_row([str(uuid.uuid4())[:8], nombre, "1","","","","","",""])
        st.cache_data.clear()

def guardar_empleado_nuevo(emp: dict):
    ws = get_ws("empleados")
    if not ws.get_all_values():
        ws.append_row(COLS_EMP)
    ws.append_row([str(uuid.uuid4())[:8], emp.get("nombre",""), "1" if emp.get("activo",True) else "0",
                   emp.get("telefono",""), emp.get("dni",""), emp.get("zona",""),
                   emp.get("vehiculo",""), emp.get("patente",""), emp.get("observaciones","")])
    st.cache_data.clear()

def actualizar_empleado(emp_id: str, emp: dict):
    ws = get_ws("empleados")
    vals = ws.get_all_values()
    if not vals: return
    for i, row in enumerate(vals[1:], start=2):
        if len(row) > 0 and row[0] == emp_id:
            ws.update(f"A{i}:I{i}", [[emp_id, emp.get("nombre",""),
                "1" if emp.get("activo",True) else "0",
                emp.get("telefono",""), emp.get("dni",""), emp.get("zona",""),
                emp.get("vehiculo",""), emp.get("patente",""), emp.get("observaciones","")]])
            st.cache_data.clear()
            return


# ══════════════════════════════════════════════════════════════════
# DATOS: REGISTROS (Horarios)
# ══════════════════════════════════════════════════════════════════
COLS_REG = ["id","empleado_id","nombre","fecha","hora_entrada","hora_salida",
            "horas_trabajadas","diferencia","inicio_ruta","fin_ruta","cargado_por","detalle"]

def cargar_registros() -> list:
    ws = get_ws("registros")
    records = ws.get_all_records()
    result = []
    for r in records:
        try:
            result.append({
                "id": str(r.get("id","")),
                "empleado_id": str(r.get("empleado_id","")),
                "nombre": r.get("nombre",""),
                "fecha": r.get("fecha",""),
                "hora_entrada": r.get("hora_entrada",""),
                "hora_salida": r.get("hora_salida",""),
                "horas_trabajadas": float(r.get("horas_trabajadas",0) or 0),
                "diferencia": float(r.get("diferencia",0) or 0),
                "inicio_ruta": r.get("inicio_ruta",""),
                "fin_ruta": r.get("fin_ruta",""),
                "cargado_por": r.get("cargado_por",""),
                "detalle": r.get("detalle",""),
            })
        except Exception:
            continue
    return result

def _init_registros_ws():
    ws = get_ws("registros")
    if not ws.get_all_values():
        ws.append_row(COLS_REG)
        st.cache_data.clear()

def guardar_registro(reg: dict):
    ws = get_ws("registros")
    if not ws.get_all_values():
        ws.append_row(COLS_REG)
    registros = cargar_registros()
    for r in registros:
        if r["empleado_id"] == reg["empleado_id"] and r["fecha"] == reg["fecha"]:
            return False, "Ya existe un registro para ese técnico en esa fecha."
    ws.append_row([str(uuid.uuid4())[:8], reg["empleado_id"], reg["nombre"], reg["fecha"],
                   reg["hora_entrada"], reg["hora_salida"], reg["horas_trabajadas"], reg["diferencia"],
                   reg.get("inicio_ruta",""), reg.get("fin_ruta",""),
                   reg.get("cargado_por",""), reg.get("detalle","")])
    st.cache_data.clear()
    return True, "OK"

def actualizar_registro(reg_id: str, reg: dict):
    ws = get_ws("registros")
    vals = ws.get_all_values()
    if not vals: return False
    for i, row in enumerate(vals[1:], start=2):
        if len(row) > 0 and row[0] == reg_id:
            ws.update(f"A{i}:L{i}", [[reg_id, reg["empleado_id"], reg["nombre"], reg["fecha"],
                reg["hora_entrada"], reg["hora_salida"], reg["horas_trabajadas"], reg["diferencia"],
                reg.get("inicio_ruta",""), reg.get("fin_ruta",""),
                reg.get("cargado_por",""), reg.get("detalle","")]])
            st.cache_data.clear()
            return True
    return False

def eliminar_registro(reg_id: str):
    ws = get_ws("registros")
    vals = ws.get_all_values()
    if not vals: return
    for i, row in enumerate(vals[1:], start=2):
        if len(row) > 0 and row[0] == reg_id:
            ws.delete_rows(i)
            st.cache_data.clear()
            return


# ══════════════════════════════════════════════════════════════════
# DATOS: SERVICIOS
# ══════════════════════════════════════════════════════════════════
COLS_SERV = ["id","fecha","equipo","patente_equipo","hora","cliente","servicio",
             "patente","detalles","estado","observaciones","cargado_por"]

def cargar_servicios() -> list:
    ws = get_ws("servicios")
    records = ws.get_all_records()
    result = []
    for r in records:
        try:
            result.append({
                "id": str(r.get("id","")),
                "fecha": r.get("fecha",""),
                "equipo": r.get("equipo",""),
                "patente_equipo": r.get("patente_equipo",""),
                "hora": r.get("hora",""),
                "cliente": r.get("cliente",""),
                "servicio": r.get("servicio",""),
                "patente": r.get("patente",""),
                "detalles": r.get("detalles",""),
                "estado": r.get("estado","PENDIENTE"),
                "observaciones": r.get("observaciones",""),
                "cargado_por": r.get("cargado_por",""),
            })
        except Exception:
            continue
    return result

def _init_servicios_ws():
    ws = get_ws("servicios")
    if not ws.get_all_values():
        ws.append_row(COLS_SERV)
        st.cache_data.clear()

def guardar_servicio(s: dict):
    ws = get_ws("servicios")
    if not ws.get_all_values():
        ws.append_row(COLS_SERV)
    ws.append_row([str(uuid.uuid4())[:8], s["fecha"], s["equipo"], s["patente_equipo"],
                   s["hora"], s["cliente"], s["servicio"], s["patente"],
                   s.get("detalles",""), s.get("estado","PENDIENTE"),
                   s.get("observaciones",""), s.get("cargado_por","")])
    st.cache_data.clear()

def actualizar_estado_servicio(serv_id: str, nuevo_estado: str):
    ws = get_ws("servicios")
    vals = ws.get_all_values()
    if not vals: return False
    headers = vals[0]
    try:
        estado_col = headers.index("estado") + 1
    except ValueError:
        return False
    for i, row in enumerate(vals[1:], start=2):
        if len(row) > 0 and row[0] == serv_id:
            ws.update_cell(i, estado_col, nuevo_estado)
            st.cache_data.clear()
            return True
    return False

def actualizar_servicio_completo(serv_id: str, s: dict):
    ws = get_ws("servicios")
    vals = ws.get_all_values()
    if not vals: return False
    for i, row in enumerate(vals[1:], start=2):
        if len(row) > 0 and row[0] == serv_id:
            ws.update(f"A{i}:L{i}", [[serv_id, s["fecha"], s["equipo"], s["patente_equipo"],
                s["hora"], s["cliente"], s["servicio"], s["patente"],
                s.get("detalles",""), s.get("estado","PENDIENTE"),
                s.get("observaciones",""), s.get("cargado_por","")]])
            st.cache_data.clear()
            return True
    return False

def eliminar_servicio(serv_id: str):
    ws = get_ws("servicios")
    vals = ws.get_all_values()
    if not vals: return
    for i, row in enumerate(vals[1:], start=2):
        if len(row) > 0 and row[0] == serv_id:
            ws.delete_rows(i)
            st.cache_data.clear()
            return


# ══════════════════════════════════════════════════════════════════
# DATOS: SERVICIOS INTERIOR
# ══════════════════════════════════════════════════════════════════
COLS_INT = ["id","fecha","distrito","localidad","unidad","tipo_vehiculo",
            "tipo_servicio","tecnico_taller","estado","detalle","cargado_por"]

def cargar_servicios_interior() -> list:
    ws = get_ws("servicios_interior")
    records = ws.get_all_records()
    result = []
    for r in records:
        try:
            result.append({
                "id": str(r.get("id","")),
                "fecha": r.get("fecha",""),
                "distrito": r.get("distrito",""),
                "localidad": r.get("localidad",""),
                "unidad": r.get("unidad",""),
                "tipo_vehiculo": r.get("tipo_vehiculo",""),
                "tipo_servicio": r.get("tipo_servicio",""),
                "tecnico_taller": r.get("tecnico_taller",""),
                "estado": r.get("estado","PENDIENTE"),
                "detalle": r.get("detalle",""),
                "cargado_por": r.get("cargado_por",""),
            })
        except Exception:
            continue
    return result

def _init_interior_ws():
    ws = get_ws("servicios_interior")
    if not ws.get_all_values():
        ws.append_row(COLS_INT)
        st.cache_data.clear()

def guardar_interior(s: dict):
    ws = get_ws("servicios_interior")
    if not ws.get_all_values():
        ws.append_row(COLS_INT)
    ws.append_row([str(uuid.uuid4())[:8], s["fecha"], s["distrito"], s["localidad"],
                   s["unidad"], s["tipo_vehiculo"], s["tipo_servicio"], s["tecnico_taller"],
                   s.get("estado","PENDIENTE"), s.get("detalle",""), s.get("cargado_por","")])
    st.cache_data.clear()

def actualizar_estado_interior(int_id: str, nuevo_estado: str):
    ws = get_ws("servicios_interior")
    vals = ws.get_all_values()
    if not vals: return False
    headers = vals[0]
    try:
        estado_col = headers.index("estado") + 1
    except ValueError:
        return False
    for i, row in enumerate(vals[1:], start=2):
        if len(row) > 0 and row[0] == int_id:
            ws.update_cell(i, estado_col, nuevo_estado)
            st.cache_data.clear()
            return True
    return False

def eliminar_interior(int_id: str):
    ws = get_ws("servicios_interior")
    vals = ws.get_all_values()
    if not vals: return
    for i, row in enumerate(vals[1:], start=2):
        if len(row) > 0 and row[0] == int_id:
            ws.delete_rows(i)
            st.cache_data.clear()
            return


# ══════════════════════════════════════════════════════════════════
# DATOS: STOCK
# ══════════════════════════════════════════════════════════════════
COLS_STOCK = ["id","codigo","producto","categoria","stock_actual","control_observacion"]
COLS_MOV = ["id","tipo","producto_id","producto","cantidad","ubicacion","fecha","cargado_por","observacion"]

UBICACIONES_STOCK = ["Oficina","Equipo 1","Equipo 2","Taller","Mendoza","Neuquén","Córdoba"]

def cargar_stock() -> list:
    ws = get_ws("stock_productos")
    records = ws.get_all_records()
    result = []
    for r in records:
        try:
            result.append({
                "id": str(r.get("id","")),
                "codigo": r.get("codigo",""),
                "producto": r.get("producto",""),
                "categoria": r.get("categoria",""),
                "stock_actual": int(r.get("stock_actual",0) or 0),
                "control_observacion": r.get("control_observacion",""),
            })
        except Exception:
            continue
    return result

def _init_stock_ws():
    ws_p = get_ws("stock_productos")
    if not ws_p.get_all_values():
        ws_p.append_row(COLS_STOCK)
    ws_m = get_ws("stock_movimientos")
    if not ws_m.get_all_values():
        ws_m.append_row(COLS_MOV)
    st.cache_data.clear()

def guardar_producto(p: dict):
    ws = get_ws("stock_productos")
    if not ws.get_all_values():
        ws.append_row(COLS_STOCK)
    ws.append_row([str(uuid.uuid4())[:8], p["codigo"], p["producto"],
                   p.get("categoria",""), p.get("stock_actual",0), p.get("control_observacion","")])
    st.cache_data.clear()

def registrar_movimiento(mov: dict):
    """Registra entrada/salida y actualiza stock_actual del producto."""
    ws_m = get_ws("stock_movimientos")
    if not ws_m.get_all_values():
        ws_m.append_row(COLS_MOV)
    ws_m.append_row([str(uuid.uuid4())[:8], mov["tipo"], mov["producto_id"], mov["producto"],
                     mov["cantidad"], mov["ubicacion"], mov["fecha"],
                     mov.get("cargado_por",""), mov.get("observacion","")])
    # Actualizar stock
    ws_p = get_ws("stock_productos")
    vals = ws_p.get_all_values()
    if not vals: return
    headers = vals[0]
    try:
        stock_col = headers.index("stock_actual") + 1
    except ValueError:
        return
    for i, row in enumerate(vals[1:], start=2):
        if len(row) > 0 and row[0] == mov["producto_id"]:
            try:
                actual = int(row[stock_col-1] or 0)
            except Exception:
                actual = 0
            delta = mov["cantidad"] if mov["tipo"] == "ENTRADA" else -mov["cantidad"]
            ws_p.update_cell(i, stock_col, actual + delta)
            st.cache_data.clear()
            return

def cargar_movimientos() -> list:
    ws = get_ws("stock_movimientos")
    records = ws.get_all_records()
    result = []
    for r in records:
        try:
            result.append({
                "id": str(r.get("id","")),
                "tipo": r.get("tipo",""),
                "producto_id": str(r.get("producto_id","")),
                "producto": r.get("producto",""),
                "cantidad": int(r.get("cantidad",0) or 0),
                "ubicacion": r.get("ubicacion",""),
                "fecha": r.get("fecha",""),
                "cargado_por": r.get("cargado_por",""),
                "observacion": r.get("observacion",""),
            })
        except Exception:
            continue
    return result


# ══════════════════════════════════════════════════════════════════
# SIDEBAR con módulos desplegables
# ══════════════════════════════════════════════════════════════════
MODULOS = [
    {
        "id": "horarios",
        "icon": "🕐",
        "label": "Horarios",
        "color": "#1e3a8a",
        "subs": [
            ("registro", "➕  Registro"),
            ("historial", "📋  Historial"),
            ("resumen", "📊  Resumen + Excel"),
            ("estadisticas", "📈  Estadísticas"),
            ("tecnicos", "👷  Técnicos"),
        ]
    },
    {
        "id": "servicios",
        "icon": "🔧",
        "label": "Servicios",
        "color": "#0f766e",
        "subs": [
            ("serv_cargar", "➕  Cargar servicio"),
            ("serv_lista", "📋  Detalle / Lista"),
        ]
    },
    {
        "id": "interior",
        "icon": "🗺️",
        "label": "Servicios Interior",
        "color": "#7c3aed",
        "subs": [
            ("int_cargar", "➕  Cargar servicio"),
            ("int_lista", "📋  Detalle / Lista"),
        ]
    },
    {
        "id": "stock",
        "icon": "📦",
        "label": "Stock",
        "color": "#b45309",
        "subs": [
            ("stock_actual", "📊  Stock actual"),
            ("stock_entrada", "📥  Entrada"),
            ("stock_salida", "📤  Salida"),
            ("stock_productos", "⚙️  Productos"),
        ]
    },
    {
        "id": "reportes",
        "icon": "📑",
        "label": "Reportes",
        "color": "#be123c",
        "subs": [
            ("reporte_cruzado", "⚡  Horarios vs Servicios"),
        ]
    },
]

def render_sidebar():
    usuario = st.session_state.get("usuario","")
    pagina = st.session_state.get("pagina","registro")

    # Determinar módulo activo
    modulo_activo = "horarios"
    for mod in MODULOS:
        for sub_key, _ in mod["subs"]:
            if pagina == sub_key:
                modulo_activo = mod["id"]
                break

    with st.sidebar:
        # Header
        st.markdown(f"""
        <div style='padding:20px 16px 12px; border-bottom:1px solid #f0f0f0; margin-bottom:8px'>
            <div style='font-size:20px; font-weight:700; color:#111827'>⚙️ Panel de Control</div>
            <div style='font-size:12px; color:#9ca3af; margin-top:2px'>Usuario: {usuario}</div>
        </div>
        """, unsafe_allow_html=True)

        for mod in MODULOS:
            is_open = st.session_state.get(f"mod_open_{mod['id']}", mod["id"] == modulo_activo)
            color = mod["color"]
            icon = mod["icon"]
            label = mod["label"]

            # Encabezado del módulo (clickeable)
            bg = color if is_open else "transparent"
            txt_color = "#ffffff" if is_open else "#374151"
            arrow = "▲" if is_open else "▼"

            st.markdown(f"""
            <div style='background:{bg}; color:{txt_color}; border-radius:8px;
                        padding:9px 12px; margin:2px 0; cursor:pointer;
                        display:flex; justify-content:space-between; align-items:center;
                        font-weight:600; font-size:14px;'>
                <span>{icon} {label}</span>
                <span style='font-size:10px; opacity:0.7'>{arrow}</span>
            </div>
            """, unsafe_allow_html=True)

            # Botón invisible para toggle
            if st.button(f"{icon} {label}", key=f"toggle_{mod['id']}", use_container_width=True):
                st.session_state[f"mod_open_{mod['id']}"] = not is_open
                st.rerun()

            # Submenús
            if is_open:
                st.markdown(f"""
                <div style='margin-left:12px; border-left:2px solid {color}40; padding-left:8px; margin-bottom:4px'>
                """, unsafe_allow_html=True)
                for sub_key, sub_label in mod["subs"]:
                    activo = pagina == sub_key
                    btn_type = "primary" if activo else "secondary"
                    if st.button(sub_label, key=f"nav_{sub_key}", use_container_width=True, type=btn_type):
                        st.session_state["pagina"] = sub_key
                        st.rerun()
                st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("---")
        if st.button("🚪 Cerrar sesión", use_container_width=True):
            st.session_state["usuario"] = None
            st.session_state["pagina"] = "registro"
            st.rerun()


# ══════════════════════════════════════════════════════════════════
# PÁGINA: LOGIN
# ══════════════════════════════════════════════════════════════════
def pagina_login():
    st.markdown("""
    <div style='text-align:center; padding:60px 0 20px'>
        <span style='font-size:48px'>⚙️</span>
        <h1 style='font-size:28px; font-weight:700; margin:8px 0 4px'>Panel de Control</h1>
        <p style='color:#6b7280; font-size:14px'>Ingresá para continuar</p>
    </div>
    """, unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1,1.2,1])
    with col2:
        with st.form("login_form"):
            usuario = st.selectbox("Usuario", ["Alejo","Martin"])
            password = st.text_input("Contraseña", type="password", placeholder="••••••••")
            if st.form_submit_button("Ingresar", use_container_width=True):
                if validar_usuario(usuario, password):
                    st.session_state["usuario"] = usuario
                    st.session_state["pagina"] = "registro"
                    st.rerun()
                else:
                    st.error("Usuario o contraseña incorrectos.")


# ══════════════════════════════════════════════════════════════════
# HORARIOS: Registro
# ══════════════════════════════════════════════════════════════════
def pagina_registro():
    st.markdown("### ➕ Cargar jornada")
    empleados = cargar_empleados(solo_activos=True)
    if not empleados:
        st.warning("No hay técnicos activos.")
        return
    nombres = [e["nombre"] for e in empleados]
    emp_map = {e["nombre"]: e["id"] for e in empleados}

    col1, col2 = st.columns(2)
    with col1:
        tecnico = st.selectbox("Técnico", nombres)
    with col2:
        fecha = st.date_input("Fecha", value=date.today())

    col3, col4 = st.columns(2)
    with col3:
        entrada = st.text_input("Hora entrada (HH:MM)", placeholder="08:00")
    with col4:
        salida = st.text_input("Hora salida (HH:MM)", placeholder="16:00")

    col5, col6 = st.columns(2)
    with col5:
        inicio_ruta = st.text_input("Inicio de ruta", placeholder="Casa / Oficina / Depósito")
    with col6:
        fin_ruta = st.text_input("Fin de ruta", placeholder="Casa / Oficina / Depósito")

    detalle = st.text_area("Detalle / Observaciones", placeholder="Observaciones del día...")

    if st.button("💾 Guardar registro", use_container_width=True):
        if not entrada or not salida:
            st.error("Completá la hora de entrada y salida.")
        else:
            try:
                trabajado, diferencia = calcular_horas(entrada, salida)
                if trabajado is None:
                    st.error("La hora de salida debe ser mayor a la de entrada.")
                else:
                    ok, msg = guardar_registro({
                        "empleado_id": emp_map[tecnico],
                        "nombre": tecnico,
                        "fecha": fecha.strftime("%Y-%m-%d"),
                        "hora_entrada": entrada.strip(),
                        "hora_salida": salida.strip(),
                        "horas_trabajadas": round(trabajado, 4),
                        "diferencia": round(diferencia, 4),
                        "inicio_ruta": inicio_ruta,
                        "fin_ruta": fin_ruta,
                        "cargado_por": st.session_state.get("usuario",""),
                        "detalle": detalle,
                    })
                    if ok:
                        st.success(f"✅ Guardado — Trabajado: {decimal_a_hhmm(trabajado)} | Balance: {decimal_a_hhmm(diferencia)}")
                    else:
                        st.error(msg)
            except ValueError:
                st.error("Formato de hora inválido. Usá HH:MM (ej: 08:00).")
    st.caption("💡 Si ya existe un registro para esa fecha, editalo desde Historial.")


# ══════════════════════════════════════════════════════════════════
# HORARIOS: Historial
# ══════════════════════════════════════════════════════════════════
def pagina_historial():
    st.markdown("### 📋 Historial de registros")
    registros = cargar_registros()
    empleados = cargar_empleados(solo_activos=False)
    nombres_todos = ["Todos"] + [e["nombre"] for e in empleados]

    col1, col2, col3 = st.columns(3)
    with col1:
        filtro_emp = st.selectbox("Técnico", nombres_todos)
    with col2:
        filtro_desde = st.date_input("Desde", value=date.today().replace(day=1))
    with col3:
        filtro_hasta = st.date_input("Hasta", value=date.today())

    filtrados = registros
    if filtro_emp != "Todos":
        filtrados = [r for r in filtrados if r["nombre"] == filtro_emp]
    filtrados = [r for r in filtrados if filtro_desde.strftime("%Y-%m-%d") <= r["fecha"] <= filtro_hasta.strftime("%Y-%m-%d")]
    filtrados = sorted(filtrados, key=lambda x: x["fecha"], reverse=True)

    if not filtrados:
        st.info("No hay registros con ese filtro.")
        return

    df = pd.DataFrame([{
        "ID": r["id"],
        "Técnico": r["nombre"],
        "Fecha": r["fecha"],
        "Entrada": r["hora_entrada"],
        "Salida": r["hora_salida"],
        "Inicio Ruta": r["inicio_ruta"],
        "Fin Ruta": r["fin_ruta"],
        "Trabajado": decimal_a_hhmm(r["horas_trabajadas"]),
        "Balance": decimal_a_hhmm(r["diferencia"]),
        "Detalle": r["detalle"],
        "Cargado por": r["cargado_por"],
    } for r in filtrados])

    st.dataframe(df.drop(columns=["ID"]), use_container_width=True, hide_index=True)

    st.markdown("---")
    st.markdown("**Editar o eliminar un registro**")
    opciones = [f"{r['fecha']} | {r['nombre']} | {r['hora_entrada']}→{r['hora_salida']}" for r in filtrados]
    sel = st.selectbox("Seleccioná un registro", opciones)
    idx_sel = opciones.index(sel)
    reg_sel = filtrados[idx_sel]

    col_e, col_d = st.columns(2)
    with col_e:
        if st.button("✏️ Editar seleccionado", use_container_width=True):
            st.session_state["editar_reg"] = reg_sel.copy()
    with col_d:
        if st.button("🗑️ Eliminar seleccionado", use_container_width=True):
            eliminar_registro(reg_sel["id"])
            st.success("Registro eliminado.")
            st.rerun()

    if "editar_reg" in st.session_state:
        g = st.session_state["editar_reg"]
        st.markdown(f"---\n**✏️ Editando — {g['nombre']} | {g['fecha']}**")
        with st.form("form_editar_reg"):
            c1, c2 = st.columns(2)
            with c1:
                e_fecha = st.date_input("Fecha", value=datetime.strptime(g["fecha"], "%Y-%m-%d").date())
                e_entrada = st.text_input("Entrada (HH:MM)", value=g["hora_entrada"])
                e_inicio = st.text_input("Inicio de ruta", value=g.get("inicio_ruta",""))
            with c2:
                e_salida = st.text_input("Salida (HH:MM)", value=g["hora_salida"])
                e_fin = st.text_input("Fin de ruta", value=g.get("fin_ruta",""))
            e_detalle = st.text_area("Detalle", value=g.get("detalle",""))
            if st.form_submit_button("💾 Guardar cambios", use_container_width=True):
                try:
                    trabajado, diferencia = calcular_horas(e_entrada, e_salida)
                    if trabajado is None:
                        st.error("La hora de salida debe ser mayor a la de entrada.")
                    else:
                        ok = actualizar_registro(g["id"], {
                            "empleado_id": g["empleado_id"], "nombre": g["nombre"],
                            "fecha": e_fecha.strftime("%Y-%m-%d"),
                            "hora_entrada": e_entrada.strip(), "hora_salida": e_salida.strip(),
                            "horas_trabajadas": round(trabajado,4), "diferencia": round(diferencia,4),
                            "inicio_ruta": e_inicio, "fin_ruta": e_fin,
                            "cargado_por": g.get("cargado_por",""), "detalle": e_detalle,
                        })
                        if ok:
                            del st.session_state["editar_reg"]
                            st.success("Registro actualizado.")
                            st.rerun()
                except ValueError:
                    st.error("Formato de hora inválido.")


# ══════════════════════════════════════════════════════════════════
# HORARIOS: Resumen + Excel
# ══════════════════════════════════════════════════════════════════
def pagina_resumen():
    st.markdown("### 📊 Resumen mensual + Excel")
    col1, col2 = st.columns(2)
    with col1:
        mes_sel = st.selectbox("Mes", MESES, index=date.today().month - 1)
    with col2:
        anio_sel = st.selectbox("Año", [2024,2025,2026,2027], index=2)

    mes_num = str(MESES.index(mes_sel)+1).zfill(2)
    anio_str = str(anio_sel)
    registros = cargar_registros()
    filtrados = [r for r in registros if r["fecha"].startswith(f"{anio_str}-{mes_num}")]

    if not filtrados:
        st.info(f"No hay registros para {mes_sel} {anio_sel}.")
        return

    resumen = {}
    for r in filtrados:
        n = r["nombre"]
        if n not in resumen:
            resumen[n] = {"dias":0,"trabajado":0.0,"balance":0.0}
        resumen[n]["dias"] += 1
        resumen[n]["trabajado"] += r["horas_trabajadas"]
        resumen[n]["balance"] += r["diferencia"]

    df_res = pd.DataFrame([{
        "Técnico": n, "Días": v["dias"],
        "Trabajado": decimal_a_hhmm(v["trabajado"]),
        "Esperado": decimal_a_hhmm(v["dias"]*JORNADA_BASE),
        "Balance": decimal_a_hhmm(v["balance"]),
    } for n,v in sorted(resumen.items())])
    st.dataframe(df_res, use_container_width=True, hide_index=True)

    total_dias = sum(v["dias"] for v in resumen.values())
    total_trab = sum(v["trabajado"] for v in resumen.values())
    total_bal = sum(v["balance"] for v in resumen.values())
    m1,m2,m3 = st.columns(3)
    m1.metric("📅 Total jornadas", total_dias)
    m2.metric("⏱️ Total trabajado", decimal_a_hhmm(total_trab))
    m3.metric("⚖️ Balance global", decimal_a_hhmm(total_bal))

    st.markdown("---")
    if st.button("📥 Exportar Excel", use_container_width=True):
        wb = Workbook()
        azul = PatternFill("solid", fgColor="1E3A8A")
        ws1 = wb.active
        ws1.title = "Resumen"
        headers1 = ["Técnico","Días","Trabajado","Esperado","Balance"]
        ws1.append(headers1)
        for c in ws1[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = azul; c.alignment = Alignment(horizontal="center")
        for _, row_data in df_res.iterrows():
            ws1.append(list(row_data))

        ws2 = wb.create_sheet("Detalle")
        headers2 = ["Técnico","Fecha","Entrada","Salida","Inicio Ruta","Fin Ruta","Trabajado","Balance","Detalle"]
        ws2.append(headers2)
        for c in ws2[1]:
            c.font = Font(bold=True, color="FFFFFF"); c.fill = azul; c.alignment = Alignment(horizontal="center")
        for r in sorted(filtrados, key=lambda x: (x["nombre"],x["fecha"])):
            ws2.append([r["nombre"],r["fecha"],r["hora_entrada"],r["hora_salida"],
                        r.get("inicio_ruta",""),r.get("fin_ruta",""),
                        decimal_a_hhmm(r["horas_trabajadas"]),decimal_a_hhmm(r["diferencia"]),r.get("detalle","")])

        for sheet in (ws1,ws2):
            for col in sheet.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                sheet.column_dimensions[col[0].column_letter].width = min(max_len+4, 40)

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        st.download_button(f"⬇️ Descargar Resumen_{mes_num}_{anio_str}.xlsx", buf,
                           f"Resumen_{mes_num}_{anio_str}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)


# ══════════════════════════════════════════════════════════════════
# HORARIOS: Estadísticas
# ══════════════════════════════════════════════════════════════════
def pagina_estadisticas():
    st.markdown("### 📈 Estadísticas")
    registros = cargar_registros()
    if not registros:
        st.info("No hay registros para mostrar estadísticas.")
        return

    empleados = cargar_empleados(solo_activos=False)
    nombres_todos = ["Todos"] + [e["nombre"] for e in empleados]
    col1, col2 = st.columns(2)
    with col1:
        filtro_emp = st.selectbox("Técnico", nombres_todos, key="est_emp")
    with col2:
        anio_sel = st.selectbox("Año", [2024,2025,2026,2027], index=2, key="est_anio")

    filtrados = [r for r in registros if r["fecha"].startswith(str(anio_sel))]
    if filtro_emp != "Todos":
        filtrados = [r for r in filtrados if r["nombre"] == filtro_emp]
    if not filtrados:
        st.info("No hay datos con ese filtro.")
        return

    total_dias = len(filtrados)
    total_horas = sum(r["horas_trabajadas"] for r in filtrados)
    total_balance = sum(r["diferencia"] for r in filtrados)
    promedio_diario = total_horas / total_dias if total_dias else 0

    m1,m2,m3,m4 = st.columns(4)
    m1.metric("📅 Jornadas", total_dias)
    m2.metric("⏱️ Total horas", decimal_a_hhmm(total_horas))
    m3.metric("📊 Promedio diario", decimal_a_hhmm(promedio_diario))
    m4.metric("⚖️ Balance acumulado", decimal_a_hhmm(total_balance))

    st.markdown("---")
    st.markdown("#### Horas por mes")
    por_mes = {}
    for r in filtrados:
        mes_num = int(r["fecha"][5:7])
        nm = MESES[mes_num-1]
        if nm not in por_mes:
            por_mes[nm] = {"trabajado":0.0,"balance":0.0,"dias":0}
        por_mes[nm]["trabajado"] += r["horas_trabajadas"]
        por_mes[nm]["balance"] += r["diferencia"]
        por_mes[nm]["dias"] += 1
    por_mes_ord = {MESES[i]: por_mes[MESES[i]] for i in range(12) if MESES[i] in por_mes}
    df_mes = pd.DataFrame([{"Mes":k,"Días":v["dias"],"Horas trabajadas":round(v["trabajado"],2),"Balance (h)":round(v["balance"],2)} for k,v in por_mes_ord.items()])
    st.bar_chart(df_mes.set_index("Mes")[["Horas trabajadas"]])
    st.dataframe(df_mes, use_container_width=True, hide_index=True)

    if filtro_emp == "Todos":
        st.markdown("---")
        st.markdown("#### Ranking por técnico")
        por_tec = {}
        for r in filtrados:
            n = r["nombre"]
            if n not in por_tec:
                por_tec[n] = {"trabajado":0.0,"dias":0,"balance":0.0}
            por_tec[n]["trabajado"] += r["horas_trabajadas"]
            por_tec[n]["dias"] += 1
            por_tec[n]["balance"] += r["diferencia"]
        df_tec = pd.DataFrame([{"Técnico":n,"Jornadas":v["dias"],"Total horas":round(v["trabajado"],2),"Balance (h)":round(v["balance"],2)} for n,v in sorted(por_tec.items(), key=lambda x: -x[1]["trabajado"])])
        st.bar_chart(df_tec.set_index("Técnico")[["Total horas"]])
        st.dataframe(df_tec, use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════
# HORARIOS: Técnicos
# ══════════════════════════════════════════════════════════════════
def pagina_tecnicos():
    st.markdown("### 👷 Técnicos")
    tab1, tab2 = st.tabs(["👥 Ver / Editar", "➕ Nuevo técnico"])

    with tab1:
        empleados = cargar_empleados(solo_activos=False)
        if not empleados:
            st.info("No hay técnicos cargados todavía.")
        else:
            for emp in empleados:
                estado = "🟢 Activo" if emp["activo"] else "🔴 Inactivo"
                with st.expander(f"**{emp['nombre']}** — {estado}", expanded=False):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.markdown(f"📞 **Tel:** {emp['telefono'] or '—'}")
                        st.markdown(f"🪪 **DNI:** {emp['dni'] or '—'}")
                        st.markdown(f"📍 **Zona:** {emp['zona'] or '—'}")
                    with col2:
                        st.markdown(f"🚗 **Vehículo:** {emp['vehiculo'] or '—'}")
                        st.markdown(f"🔖 **Patente:** {emp['patente'] or '—'}")
                    if emp['observaciones']:
                        st.markdown(f"📝 **Obs:** {emp['observaciones']}")
                    if st.button(f"✏️ Editar {emp['nombre']}", key=f"edit_{emp['id']}"):
                        st.session_state["editar_emp"] = emp.copy()

        if "editar_emp" in st.session_state:
            e = st.session_state["editar_emp"]
            st.markdown(f"---\n**✏️ Editando — {e['nombre']}**")
            with st.form("form_edit_emp"):
                c1, c2 = st.columns(2)
                with c1:
                    e_nombre = st.text_input("Nombre", value=e["nombre"])
                    e_tel = st.text_input("Teléfono", value=e.get("telefono",""))
                    e_dni = st.text_input("DNI", value=e.get("dni",""))
                    e_zona = st.text_input("Zona", value=e.get("zona",""))
                with c2:
                    e_veh = st.text_input("Vehículo", value=e.get("vehiculo",""))
                    e_pat = st.text_input("Patente", value=e.get("patente",""))
                    e_activo = st.checkbox("Activo", value=e.get("activo",True))
                e_obs = st.text_area("Observaciones", value=e.get("observaciones",""))
                if st.form_submit_button("💾 Guardar cambios", use_container_width=True):
                    if not e_nombre:
                        st.error("El nombre no puede estar vacío.")
                    else:
                        actualizar_empleado(e["id"], {"nombre":e_nombre,"activo":e_activo,"telefono":e_tel,
                                                       "dni":e_dni,"zona":e_zona,"vehiculo":e_veh,
                                                       "patente":e_pat,"observaciones":e_obs})
                        del st.session_state["editar_emp"]
                        st.success("Técnico actualizado.")
                        st.rerun()

    with tab2:
        st.markdown("#### Nuevo técnico")
        with st.form("form_nuevo_emp", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                n_nombre = st.text_input("Nombre *", placeholder="Maxi")
                n_tel = st.text_input("Teléfono", placeholder="11 1234-5678")
                n_dni = st.text_input("DNI", placeholder="12345678")
                n_zona = st.text_input("Zona", placeholder="CABA / Zona Oeste")
            with c2:
                n_veh = st.text_input("Vehículo", placeholder="Partner / Moto")
                n_pat = st.text_input("Patente", placeholder="AB123CD")
                n_activo = st.checkbox("Activo", value=True)
            n_obs = st.text_area("Observaciones")
            if st.form_submit_button("💾 Guardar técnico", use_container_width=True):
                if not n_nombre:
                    st.error("El nombre es obligatorio.")
                else:
                    guardar_empleado_nuevo({"nombre":n_nombre,"activo":n_activo,"telefono":n_tel,
                                            "dni":n_dni,"zona":n_zona,"vehiculo":n_veh,
                                            "patente":n_pat,"observaciones":n_obs})
                    st.success(f"✅ Técnico '{n_nombre}' agregado.")
                    st.rerun()


# ══════════════════════════════════════════════════════════════════
# SERVICIOS: Cargar
# ══════════════════════════════════════════════════════════════════
def pagina_serv_cargar():
    st.markdown("### ➕ Cargar servicio")
    with st.form("form_serv_nuevo", clear_on_submit=True):
        col1, col2, col3 = st.columns(3)
        with col1:
            s_fecha = st.date_input("Fecha", value=date.today())
        with col2:
            s_equipo = st.selectbox("Equipo", list(EQUIPOS.keys()))
        with col3:
            s_hora = st.text_input("Hora (HH:MM)", placeholder="10:00")

        col4, col5 = st.columns(2)
        with col4:
            s_cliente = st.text_input("Cliente *", placeholder="QUILMES DISTRIBUIDORES")
        with col5:
            s_servicio = st.selectbox("Servicio", ["INSTALACION","DESINSTALACION","REVISION","MANTENIMIENTO","OTRO"])

        col6, col7 = st.columns(2)
        with col6:
            s_patente = st.text_input("Patente del vehículo", placeholder="HQT470")
        with col7:
            s_detalles = st.text_input("Detalles", placeholder="GPS / CÁMARA / etc.")

        col8, col9 = st.columns(2)
        with col8:
            s_estado = st.selectbox("Estado", ESTADOS_SERVICIO)
        with col9:
            s_obs = st.text_input("Observaciones", placeholder="Dirección, contacto, etc.")

        if st.form_submit_button("💾 Guardar servicio", use_container_width=True):
            if not s_cliente:
                st.error("El cliente es obligatorio.")
            else:
                guardar_servicio({
                    "fecha": s_fecha.strftime("%Y-%m-%d"),
                    "equipo": s_equipo,
                    "patente_equipo": EQUIPOS[s_equipo],
                    "hora": s_hora.strip(),
                    "cliente": s_cliente,
                    "servicio": s_servicio,
                    "patente": s_patente,
                    "detalles": s_detalles,
                    "estado": s_estado,
                    "observaciones": s_obs,
                    "cargado_por": st.session_state.get("usuario",""),
                })
                st.success("✅ Servicio guardado.")


# ══════════════════════════════════════════════════════════════════
# SERVICIOS: Lista / Detalle
# ══════════════════════════════════════════════════════════════════
def estado_color(estado: str) -> str:
    colores = {
        "REALIZADO": "🟢",
        "CONFIRMADO": "🔵",
        "PENDIENTE": "🟡",
        "SUSPENDIDO": "🔴",
        "REPROGRAMADO": "🟠",
    }
    return colores.get(estado.upper(), "⚪")

def pagina_serv_lista():
    st.markdown("### 📋 Detalle / Lista de servicios")
    servicios = cargar_servicios()

    # Filtros
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        filtro_equipo = st.selectbox("Equipo", ["Todos"] + list(EQUIPOS.keys()))
    with col2:
        filtro_desde = st.date_input("Desde", value=date.today().replace(day=1))
    with col3:
        filtro_hasta = st.date_input("Hasta", value=date.today())
    with col4:
        filtro_estado = st.selectbox("Estado", ["Todos"] + ESTADOS_SERVICIO)

    filtro_cliente = st.text_input("Buscar cliente", placeholder="Escribí parte del nombre...")

    filtrados = servicios
    if filtro_equipo != "Todos":
        filtrados = [s for s in filtrados if s["equipo"] == filtro_equipo]
    if filtro_estado != "Todos":
        filtrados = [s for s in filtrados if s["estado"].upper() == filtro_estado]
    if filtro_cliente:
        filtrados = [s for s in filtrados if filtro_cliente.lower() in s["cliente"].lower()]
    filtrados = [s for s in filtrados if filtro_desde.strftime("%Y-%m-%d") <= s["fecha"] <= filtro_hasta.strftime("%Y-%m-%d")]
    filtrados = sorted(filtrados, key=lambda x: (x["fecha"], x["hora"]), reverse=True)

    if not filtrados:
        st.info("No hay servicios con ese filtro.")
        return

    # Métricas rápidas
    m1,m2,m3,m4,m5 = st.columns(5)
    m1.metric("Total", len(filtrados))
    m2.metric("🟢 Realizados", sum(1 for s in filtrados if s["estado"].upper()=="REALIZADO"))
    m3.metric("🔵 Confirmados", sum(1 for s in filtrados if s["estado"].upper()=="CONFIRMADO"))
    m4.metric("🟡 Pendientes", sum(1 for s in filtrados if s["estado"].upper()=="PENDIENTE"))
    m5.metric("🔴 Suspendidos", sum(1 for s in filtrados if s["estado"].upper() in ["SUSPENDIDO","REPROGRAMADO"]))

    st.markdown("---")

    # Tabla editable para cambiar estado
    st.markdown("#### Actualizar estados")
    st.caption("Editá el estado directamente en la tabla y hacé click en Guardar cambios.")

    df_edit = pd.DataFrame([{
        "id": s["id"],
        "Fecha": s["fecha"],
        "Equipo": s["equipo"],
        "Hora": s["hora"],
        "Cliente": s["cliente"],
        "Servicio": s["servicio"],
        "Patente": s["patente"],
        "Detalles": s["detalles"],
        "Estado": s["estado"],
        "Observaciones": s["observaciones"],
    } for s in filtrados])

    edited_df = st.data_editor(
        df_edit.drop(columns=["id"]),
        use_container_width=True,
        hide_index=True,
        column_config={
            "Estado": st.column_config.SelectboxColumn(
                "Estado",
                options=ESTADOS_SERVICIO,
                required=True,
            ),
            "Fecha": st.column_config.TextColumn("Fecha", disabled=True),
            "Equipo": st.column_config.TextColumn("Equipo", disabled=True),
            "Hora": st.column_config.TextColumn("Hora", disabled=True),
            "Cliente": st.column_config.TextColumn("Cliente", disabled=True),
            "Servicio": st.column_config.TextColumn("Servicio", disabled=True),
            "Patente": st.column_config.TextColumn("Patente", disabled=True),
            "Detalles": st.column_config.TextColumn("Detalles", disabled=True),
            "Observaciones": st.column_config.TextColumn("Observaciones", disabled=True),
        }
    )

    if st.button("💾 Guardar cambios de estado", use_container_width=True):
        cambios = 0
        for i, row in edited_df.iterrows():
            serv_id = df_edit.iloc[i]["id"]
            estado_nuevo = row["Estado"]
            estado_original = df_edit.iloc[i]["Estado"]
            if estado_nuevo != estado_original:
                actualizar_estado_servicio(serv_id, estado_nuevo)
                cambios += 1
        if cambios > 0:
            st.success(f"✅ {cambios} estado(s) actualizado(s).")
            st.rerun()
        else:
            st.info("No hubo cambios.")

    # Eliminar
    st.markdown("---")
    st.markdown("**Eliminar un servicio**")
    opciones_del = [f"{s['fecha']} | {s['equipo']} | {s['hora']} | {s['cliente']}" for s in filtrados]
    sel_del = st.selectbox("Seleccioná un servicio", opciones_del, key="del_serv")
    if st.button("🗑️ Eliminar seleccionado", key="btn_del_serv"):
        idx_del = opciones_del.index(sel_del)
        eliminar_servicio(filtrados[idx_del]["id"])
        st.success("Servicio eliminado.")
        st.rerun()


# ══════════════════════════════════════════════════════════════════
# SERVICIOS INTERIOR: Cargar
# ══════════════════════════════════════════════════════════════════
def pagina_int_cargar():
    st.markdown("### ➕ Cargar servicio interior")
    with st.form("form_int_nuevo", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            i_fecha = st.date_input("Fecha", value=date.today())
        with col2:
            i_tecnico = st.text_input("Técnico / Taller *", placeholder="Eduardo Zarza / Vitaco / Taller")

        col3, col4 = st.columns(2)
        with col3:
            i_distrito = st.text_input("Distrito / Cliente", placeholder="La Serenísima / MANTELECTRIC")
        with col4:
            i_localidad = st.text_input("Localidad", placeholder="CD Rosario / Córdoba")

        col5, col6 = st.columns(2)
        with col5:
            i_unidad = st.text_input("Unidad / Patente", placeholder="KWN846 / AF104TN")
        with col6:
            i_tipo_veh = st.selectbox("Tipo de vehículo", ["CHASIS","PICK-UP","AUTOMOVIL","SEMI","MOTO","OTRO"])

        col7, col8 = st.columns(2)
        with col7:
            i_tipo_serv = st.selectbox("Tipo de servicio", ["Instalación","Desinstalación","Revisión","Mantenimiento","Otro"])
        with col8:
            i_estado = st.selectbox("Estado", ESTADOS_SERVICIO)

        i_detalle = st.text_area("Detalle", placeholder="GPS, corte corriente, botón de pánico...")

        if st.form_submit_button("💾 Guardar", use_container_width=True):
            if not i_tecnico:
                st.error("El técnico / taller es obligatorio.")
            else:
                guardar_interior({
                    "fecha": i_fecha.strftime("%Y-%m-%d"),
                    "distrito": i_distrito,
                    "localidad": i_localidad,
                    "unidad": i_unidad,
                    "tipo_vehiculo": i_tipo_veh,
                    "tipo_servicio": i_tipo_serv,
                    "tecnico_taller": i_tecnico,
                    "estado": i_estado,
                    "detalle": i_detalle,
                    "cargado_por": st.session_state.get("usuario",""),
                })
                st.success("✅ Servicio interior guardado.")


# ══════════════════════════════════════════════════════════════════
# SERVICIOS INTERIOR: Lista
# ══════════════════════════════════════════════════════════════════
def pagina_int_lista():
    st.markdown("### 📋 Detalle / Lista — Servicios Interior")
    servicios = cargar_servicios_interior()

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        filtro_desde = st.date_input("Desde", value=date.today().replace(day=1))
    with col2:
        filtro_hasta = st.date_input("Hasta", value=date.today())
    with col3:
        filtro_estado = st.selectbox("Estado", ["Todos"] + ESTADOS_SERVICIO)
    with col4:
        filtro_tecnico = st.text_input("Técnico / Taller", placeholder="Nombre...")

    filtro_cliente = st.text_input("Distrito / Cliente", placeholder="La Serenísima / MANTELECTRIC...")

    filtrados = servicios
    if filtro_estado != "Todos":
        filtrados = [s for s in filtrados if s["estado"].upper() == filtro_estado]
    if filtro_tecnico:
        filtrados = [s for s in filtrados if filtro_tecnico.lower() in s["tecnico_taller"].lower()]
    if filtro_cliente:
        filtrados = [s for s in filtrados if filtro_cliente.lower() in s["distrito"].lower()]
    filtrados = [s for s in filtrados if filtro_desde.strftime("%Y-%m-%d") <= s["fecha"] <= filtro_hasta.strftime("%Y-%m-%d")]
    filtrados = sorted(filtrados, key=lambda x: x["fecha"], reverse=True)

    if not filtrados:
        st.info("No hay servicios con ese filtro.")
        return

    m1,m2,m3 = st.columns(3)
    m1.metric("Total", len(filtrados))
    m2.metric("🟢 Realizados", sum(1 for s in filtrados if s["estado"].upper()=="REALIZADO"))
    m3.metric("🟡 Pendientes", sum(1 for s in filtrados if s["estado"].upper()=="PENDIENTE"))

    st.markdown("---")
    st.markdown("#### Actualizar estados")

    df_edit = pd.DataFrame([{
        "id": s["id"],
        "Fecha": s["fecha"],
        "Distrito": s["distrito"],
        "Localidad": s["localidad"],
        "Unidad": s["unidad"],
        "Tipo Vehículo": s["tipo_vehiculo"],
        "Tipo Servicio": s["tipo_servicio"],
        "Técnico / Taller": s["tecnico_taller"],
        "Estado": s["estado"],
        "Detalle": s["detalle"],
    } for s in filtrados])

    edited_df = st.data_editor(
        df_edit.drop(columns=["id"]),
        use_container_width=True,
        hide_index=True,
        column_config={
            "Estado": st.column_config.SelectboxColumn("Estado", options=ESTADOS_SERVICIO, required=True),
            "Fecha": st.column_config.TextColumn("Fecha", disabled=True),
            "Distrito": st.column_config.TextColumn("Distrito", disabled=True),
            "Localidad": st.column_config.TextColumn("Localidad", disabled=True),
            "Unidad": st.column_config.TextColumn("Unidad", disabled=True),
            "Tipo Vehículo": st.column_config.TextColumn("Tipo Vehículo", disabled=True),
            "Tipo Servicio": st.column_config.TextColumn("Tipo Servicio", disabled=True),
            "Técnico / Taller": st.column_config.TextColumn("Técnico / Taller", disabled=True),
            "Detalle": st.column_config.TextColumn("Detalle", disabled=True),
        }
    )

    if st.button("💾 Guardar cambios de estado", use_container_width=True, key="save_int"):
        cambios = 0
        for i, row in edited_df.iterrows():
            int_id = df_edit.iloc[i]["id"]
            estado_nuevo = row["Estado"]
            if estado_nuevo != df_edit.iloc[i]["Estado"]:
                actualizar_estado_interior(int_id, estado_nuevo)
                cambios += 1
        if cambios > 0:
            st.success(f"✅ {cambios} estado(s) actualizado(s).")
            st.rerun()
        else:
            st.info("No hubo cambios.")

    st.markdown("---")
    opciones_del = [f"{s['fecha']} | {s['distrito']} | {s['unidad']} | {s['tecnico_taller']}" for s in filtrados]
    sel_del = st.selectbox("Seleccioná para eliminar", opciones_del)
    if st.button("🗑️ Eliminar seleccionado", key="del_int"):
        idx_del = opciones_del.index(sel_del)
        eliminar_interior(filtrados[idx_del]["id"])
        st.success("Eliminado.")
        st.rerun()


# ══════════════════════════════════════════════════════════════════
# STOCK: Stock actual
# ══════════════════════════════════════════════════════════════════
def pagina_stock_actual():
    st.markdown("### 📊 Stock actual")
    productos = cargar_stock()
    if not productos:
        st.info("No hay productos cargados. Agregalos en la sección Productos.")
        return

    df = pd.DataFrame([{
        "Código": p["codigo"],
        "Producto": p["producto"],
        "Categoría": p["categoria"],
        "Stock": p["stock_actual"],
        "Obs": p["control_observacion"],
    } for p in productos])

    st.dataframe(df, use_container_width=True, hide_index=True)

    total = sum(p["stock_actual"] for p in productos)
    st.metric("Total unidades en stock", total)


# ══════════════════════════════════════════════════════════════════
# STOCK: Entrada
# ══════════════════════════════════════════════════════════════════
def pagina_stock_entrada():
    st.markdown("### 📥 Registrar entrada de stock")
    productos = cargar_stock()
    if not productos:
        st.warning("No hay productos. Agregalos primero en Productos.")
        return

    prod_map = {f"{p['codigo']} — {p['producto']}": p for p in productos}
    with st.form("form_entrada", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            prod_sel = st.selectbox("Producto", list(prod_map.keys()))
        with col2:
            cantidad = st.number_input("Cantidad", min_value=1, value=1)
        col3, col4 = st.columns(2)
        with col3:
            ubicacion = st.selectbox("Ubicación destino", UBICACIONES_STOCK)
        with col4:
            fecha_mov = st.date_input("Fecha", value=date.today())
        obs = st.text_input("Observación", placeholder="Proveedor, nro de remito, etc.")

        if st.form_submit_button("💾 Registrar entrada", use_container_width=True):
            p = prod_map[prod_sel]
            registrar_movimiento({
                "tipo": "ENTRADA",
                "producto_id": p["id"],
                "producto": p["producto"],
                "cantidad": cantidad,
                "ubicacion": ubicacion,
                "fecha": fecha_mov.strftime("%Y-%m-%d"),
                "cargado_por": st.session_state.get("usuario",""),
                "observacion": obs,
            })
            st.success(f"✅ Entrada de {cantidad} unidades registrada. Stock actualizado.")


# ══════════════════════════════════════════════════════════════════
# STOCK: Salida
# ══════════════════════════════════════════════════════════════════
def pagina_stock_salida():
    st.markdown("### 📤 Registrar salida de stock")
    productos = cargar_stock()
    if not productos:
        st.warning("No hay productos. Agregalos primero en Productos.")
        return

    prod_map = {f"{p['codigo']} — {p['producto']} (stock: {p['stock_actual']})": p for p in productos}
    with st.form("form_salida", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            prod_sel = st.selectbox("Producto", list(prod_map.keys()))
        with col2:
            cantidad = st.number_input("Cantidad", min_value=1, value=1)
        col3, col4 = st.columns(2)
        with col3:
            ubicacion = st.selectbox("Destino", UBICACIONES_STOCK)
        with col4:
            fecha_mov = st.date_input("Fecha", value=date.today())
        obs = st.text_input("Observación", placeholder="Número de servicio, cliente, etc.")

        if st.form_submit_button("💾 Registrar salida", use_container_width=True):
            p = prod_map[prod_sel]
            if cantidad > p["stock_actual"]:
                st.error(f"Stock insuficiente. Stock actual: {p['stock_actual']}")
            else:
                registrar_movimiento({
                    "tipo": "SALIDA",
                    "producto_id": p["id"],
                    "producto": p["producto"],
                    "cantidad": cantidad,
                    "ubicacion": ubicacion,
                    "fecha": fecha_mov.strftime("%Y-%m-%d"),
                    "cargado_por": st.session_state.get("usuario",""),
                    "observacion": obs,
                })
                st.success(f"✅ Salida de {cantidad} unidades registrada. Stock actualizado.")


# ══════════════════════════════════════════════════════════════════
# STOCK: Productos (CRUD)
# ══════════════════════════════════════════════════════════════════
def pagina_stock_productos():
    st.markdown("### ⚙️ Gestión de productos")
    tab1, tab2 = st.tabs(["📋 Ver productos", "➕ Nuevo producto"])

    with tab1:
        productos = cargar_stock()
        if not productos:
            st.info("No hay productos cargados.")
        else:
            df = pd.DataFrame([{
                "Código": p["codigo"],
                "Producto": p["producto"],
                "Categoría": p["categoria"],
                "Stock": p["stock_actual"],
                "Observación": p["control_observacion"],
            } for p in productos])
            st.dataframe(df, use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("**Historial de movimientos**")
        movimientos = cargar_movimientos()
        if movimientos:
            df_mov = pd.DataFrame([{
                "Tipo": m["tipo"],
                "Producto": m["producto"],
                "Cantidad": m["cantidad"],
                "Ubicación": m["ubicacion"],
                "Fecha": m["fecha"],
                "Cargado por": m["cargado_por"],
                "Observación": m["observacion"],
            } for m in sorted(movimientos, key=lambda x: x["fecha"], reverse=True)])
            st.dataframe(df_mov, use_container_width=True, hide_index=True)

    with tab2:
        with st.form("form_nuevo_prod", clear_on_submit=True):
            col1, col2 = st.columns(2)
            with col1:
                p_codigo = st.text_input("Código *", placeholder="D01 / I05")
                p_nombre = st.text_input("Producto *", placeholder="TRAX S40 (NUEVOS)")
            with col2:
                p_cat = st.selectbox("Categoría", ["DISPOSITIVOS","INSUMOS","HERRAMIENTAS","OTROS"])
                p_stock = st.number_input("Stock inicial", min_value=0, value=0)
            p_obs = st.text_input("Observación / Control")
            if st.form_submit_button("💾 Agregar producto", use_container_width=True):
                if not p_codigo or not p_nombre:
                    st.error("Código y producto son obligatorios.")
                else:
                    guardar_producto({"codigo":p_codigo,"producto":p_nombre,"categoria":p_cat,
                                      "stock_actual":p_stock,"control_observacion":p_obs})
                    st.success(f"✅ Producto '{p_nombre}' agregado.")
                    st.rerun()


# ══════════════════════════════════════════════════════════════════
# REPORTES: Horarios vs Servicios
# ══════════════════════════════════════════════════════════════════
def pagina_reporte_cruzado():
    st.markdown("### ⚡ Reporte cruzado — Horarios vs Servicios")
    st.caption("Compará las horas trabajadas de un técnico con la cantidad de servicios realizados por su equipo.")

    col1, col2, col3 = st.columns(3)
    with col1:
        mes_sel = st.selectbox("Mes", MESES, index=date.today().month - 1, key="rep_mes")
    with col2:
        anio_sel = st.selectbox("Año", [2024,2025,2026,2027], index=2, key="rep_anio")
    with col3:
        empleados = cargar_empleados(solo_activos=False)
        nombres = [e["nombre"] for e in empleados]
        tecnico_sel = st.selectbox("Técnico", nombres if nombres else ["—"])

    mes_num = str(MESES.index(mes_sel)+1).zfill(2)
    anio_str = str(anio_sel)
    prefijo = f"{anio_str}-{mes_num}"

    # Buscar vehículo del técnico
    emp_obj = next((e for e in empleados if e["nombre"] == tecnico_sel), None)
    patente_tec = emp_obj["patente"] if emp_obj else ""

    # Inferir equipo por patente
    equipo_tec = None
    for eq, pat in EQUIPOS.items():
        if pat.upper() == patente_tec.upper():
            equipo_tec = eq
            break

    # Registros de horarios
    registros = [r for r in cargar_registros() if r["nombre"] == tecnico_sel and r["fecha"].startswith(prefijo)]
    # Servicios del equipo
    servicios = []
    if equipo_tec:
        servicios = [s for s in cargar_servicios()
                     if s["equipo"] == equipo_tec and s["fecha"].startswith(prefijo)
                     and s["estado"].upper() == "REALIZADO"]

    st.markdown("---")

    if not registros:
        st.info(f"No hay registros de horario para {tecnico_sel} en {mes_sel} {anio_sel}.")
    
    # Métricas globales
    total_horas = sum(r["horas_trabajadas"] for r in registros)
    total_balance = sum(r["diferencia"] for r in registros)
    dias_trabajados = len(registros)
    total_servicios = len(servicios)

    m1,m2,m3,m4,m5 = st.columns(5)
    m1.metric("👷 Técnico", tecnico_sel)
    m2.metric("🚗 Equipo", equipo_tec or "No asignado")
    m3.metric("📅 Días trabajados", dias_trabajados)
    m4.metric("⏱️ Horas trabajadas", decimal_a_hhmm(total_horas))
    m5.metric("🔧 Servicios realizados", total_servicios)

    if dias_trabajados > 0 and total_servicios > 0:
        prom_servicios_dia = total_servicios / dias_trabajados
        prom_hs_servicio = total_horas / total_servicios if total_servicios else 0
        st.markdown("---")
        c1,c2,c3 = st.columns(3)
        c1.metric("📊 Servicios / día promedio", f"{prom_servicios_dia:.1f}")
        c2.metric("⏱️ Horas / servicio promedio", f"{prom_hs_servicio:.1f}h")
        c3.metric("⚖️ Balance acumulado", decimal_a_hhmm(total_balance))

    # Tabla cruzada por día
    if registros:
        st.markdown("---")
        st.markdown("#### Detalle por día")
        servicios_por_dia = {}
        for s in servicios:
            d = s["fecha"]
            servicios_por_dia[d] = servicios_por_dia.get(d, 0) + 1

        df_cruzado = pd.DataFrame([{
            "Fecha": r["fecha"],
            "Entrada": r["hora_entrada"],
            "Salida": r["hora_salida"],
            "Horas trabajadas": decimal_a_hhmm(r["horas_trabajadas"]),
            "Balance": decimal_a_hhmm(r["diferencia"]),
            "Servicios realizados": servicios_por_dia.get(r["fecha"], 0),
        } for r in sorted(registros, key=lambda x: x["fecha"])])

        st.dataframe(df_cruzado, use_container_width=True, hide_index=True)

    if not equipo_tec:
        st.warning(f"⚠️ {tecnico_sel} no tiene una patente asignada que coincida con los equipos ({', '.join(EQUIPOS.values())}). Actualizalo en Técnicos.")


# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════
def main():
    if "usuario" not in st.session_state:
        st.session_state["usuario"] = None
    if "pagina" not in st.session_state:
        st.session_state["pagina"] = "registro"

    if not st.session_state["usuario"]:
        pagina_login()
        return

    # Inicializar hojas
    _init_empleados_ws()
    _init_registros_ws()
    _init_servicios_ws()
    _init_interior_ws()
    _init_stock_ws()

    render_sidebar()

    # Routing
    pagina = st.session_state["pagina"]
    rutas = {
        "registro": pagina_registro,
        "historial": pagina_historial,
        "resumen": pagina_resumen,
        "estadisticas": pagina_estadisticas,
        "tecnicos": pagina_tecnicos,
        "serv_cargar": pagina_serv_cargar,
        "serv_lista": pagina_serv_lista,
        "int_cargar": pagina_int_cargar,
        "int_lista": pagina_int_lista,
        "stock_actual": pagina_stock_actual,
        "stock_entrada": pagina_stock_entrada,
        "stock_salida": pagina_stock_salida,
        "stock_productos": pagina_stock_productos,
        "reporte_cruzado": pagina_reporte_cruzado,
    }
    fn = rutas.get(pagina)
    if fn:
        fn()


if __name__ == "__main__":
    main()