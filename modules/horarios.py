from datetime import date, datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io
import streamlit as st

from config import MESES, JORNADA_BASE
from utils import decimal_a_hhmm, calcular_horas
from db.empleados import (
    cargar_empleados, guardar_empleado_nuevo, actualizar_empleado,
)
from db.registros import (
    cargar_registros, guardar_registro, actualizar_registro, eliminar_registro,
)


# ─────────────────────────────────────────────
# Registro
# ─────────────────────────────────────────────
def pagina_registro():
    st.markdown("### ➕ Cargar jornada")
    empleados = cargar_empleados(solo_activos=True)
    if not empleados:
        st.warning("No hay técnicos activos.")
        return

    nombres  = [e["nombre"] for e in empleados]
    emp_map  = {e["nombre"]: e["id"] for e in empleados}

    col1, col2 = st.columns(2)
    with col1:
        tecnico = st.selectbox("Técnico", nombres)
    with col2:
        fecha = st.date_input("Fecha", value=date.today())

    col3, col4 = st.columns(2)
    with col3:
        entrada = st.text_input("Hora entrada (HH:MM)", placeholder="08:00")
    with col4:
        salida = st.text_input("Hora salida (HH:MM)", placeholder="16:00")

    col5, col6 = st.columns(2)
    with col5:
        inicio_ruta = st.text_input("Inicio de ruta", placeholder="Casa / Oficina / Depósito")
    with col6:
        fin_ruta = st.text_input("Fin de ruta", placeholder="Casa / Oficina / Depósito")

    detalle = st.text_area("Detalle / Observaciones", placeholder="Observaciones del día...")

    if st.button("💾 Guardar registro", use_container_width=True):
        if not entrada or not salida:
            st.error("Completá la hora de entrada y salida.")
        else:
            trabajado, diferencia = calcular_horas(entrada, salida)
            if trabajado is None:
                st.error("Formato inválido o salida antes de entrada. Usá HH:MM (ej: 08:00).")
            else:
                ok, msg = guardar_registro({
                    "empleado_id":     emp_map[tecnico],
                    "nombre":          tecnico,
                    "fecha":           fecha.strftime("%Y-%m-%d"),
                    "hora_entrada":    entrada.strip(),
                    "hora_salida":     salida.strip(),
                    "horas_trabajadas": round(trabajado, 4),
                    "diferencia":      round(diferencia, 4),
                    "inicio_ruta":     inicio_ruta,
                    "fin_ruta":        fin_ruta,
                    "cargado_por":     st.session_state.get("usuario", ""),
                    "detalle":         detalle,
                })
                if ok:
                    st.success(f"✅ Guardado — Trabajado: {decimal_a_hhmm(trabajado)} | Balance: {decimal_a_hhmm(diferencia)}")
                else:
                    st.error(msg)

    st.caption("💡 Si ya existe un registro para esa fecha, editalo desde Historial.")


# ─────────────────────────────────────────────
# Historial
# ─────────────────────────────────────────────
def pagina_historial():
    st.markdown("### 📋 Historial de registros")
    registros = cargar_registros()
    empleados = cargar_empleados(solo_activos=False)
    nombres_todos = ["Todos"] + [e["nombre"] for e in empleados]

    col1, col2, col3 = st.columns(3)
    with col1:
        filtro_emp = st.selectbox("Técnico", nombres_todos)
    with col2:
        filtro_desde = st.date_input("Desde", value=date.today().replace(day=1))
    with col3:
        filtro_hasta = st.date_input("Hasta", value=date.today())

    filtrados = registros
    if filtro_emp != "Todos":
        filtrados = [r for r in filtrados if r["nombre"] == filtro_emp]
    filtrados = [
        r for r in filtrados
        if filtro_desde.strftime("%Y-%m-%d") <= r["fecha"] <= filtro_hasta.strftime("%Y-%m-%d")
    ]
    filtrados = sorted(filtrados, key=lambda x: x["fecha"], reverse=True)

    if not filtrados:
        st.info("No hay registros con ese filtro.")
        return

    df = pd.DataFrame([{
        "ID":          r["id"],
        "Técnico":     r["nombre"],
        "Fecha":       r["fecha"],
        "Entrada":     r["hora_entrada"],
        "Salida":      r["hora_salida"],
        "Inicio Ruta": r["inicio_ruta"],
        "Fin Ruta":    r["fin_ruta"],
        "Trabajado":   decimal_a_hhmm(r["horas_trabajadas"]),
        "Balance":     decimal_a_hhmm(r["diferencia"]),
        "Detalle":     r["detalle"],
        "Cargado por": r["cargado_por"],
    } for r in filtrados])

    st.dataframe(df.drop(columns=["ID"]), use_container_width=True, hide_index=True)

    st.markdown("---")
    st.markdown("**Editar o eliminar un registro**")
    opciones = [f"{r['fecha']} | {r['nombre']} | {r['hora_entrada']}→{r['hora_salida']}" for r in filtrados]
    sel      = st.selectbox("Seleccioná un registro", opciones)
    reg_sel  = filtrados[opciones.index(sel)]

    col_e, col_d = st.columns(2)
    with col_e:
        if st.button("✏️ Editar seleccionado", use_container_width=True):
            st.session_state["editar_reg"] = reg_sel.copy()
    with col_d:
        if st.button("🗑️ Eliminar seleccionado", use_container_width=True):
            eliminar_registro(reg_sel["id"])
            st.success("Registro eliminado.")
            st.rerun()

    if "editar_reg" in st.session_state:
        g = st.session_state["editar_reg"]
        st.markdown(f"---\n**✏️ Editando — {g['nombre']} | {g['fecha']}**")
        with st.form("form_editar_reg"):
            c1, c2 = st.columns(2)
            with c1:
                e_fecha   = st.date_input("Fecha",         value=datetime.strptime(g["fecha"], "%Y-%m-%d").date())
                e_entrada = st.text_input("Entrada (HH:MM)", value=g["hora_entrada"])
                e_inicio  = st.text_input("Inicio de ruta",  value=g.get("inicio_ruta", ""))
            with c2:
                e_salida  = st.text_input("Salida (HH:MM)", value=g["hora_salida"])
                e_fin     = st.text_input("Fin de ruta",    value=g.get("fin_ruta", ""))
            e_detalle = st.text_area("Detalle", value=g.get("detalle", ""))
            if st.form_submit_button("💾 Guardar cambios", use_container_width=True):
                trabajado, diferencia = calcular_horas(e_entrada, e_salida)
                if trabajado is None:
                    st.error("Formato inválido o salida antes de entrada.")
                else:
                    ok = actualizar_registro(g["id"], {
                        "empleado_id":     g["empleado_id"],
                        "nombre":          g["nombre"],
                        "fecha":           e_fecha.strftime("%Y-%m-%d"),
                        "hora_entrada":    e_entrada.strip(),
                        "hora_salida":     e_salida.strip(),
                        "horas_trabajadas": round(trabajado, 4),
                        "diferencia":      round(diferencia, 4),
                        "inicio_ruta":     e_inicio,
                        "fin_ruta":        e_fin,
                        "cargado_por":     g.get("cargado_por", ""),
                        "detalle":         e_detalle,
                    })
                    if ok:
                        del st.session_state["editar_reg"]
                        st.success("Registro actualizado.")
                        st.rerun()


# ─────────────────────────────────────────────
# Resumen + Excel
# ─────────────────────────────────────────────
def pagina_resumen():
    st.markdown("### 📊 Resumen mensual + Excel")
    col1, col2 = st.columns(2)
    with col1:
        mes_sel  = st.selectbox("Mes",  MESES, index=date.today().month - 1)
    with col2:
        anio_sel = st.selectbox("Año",  [2024, 2025, 2026, 2027], index=2)

    mes_num  = str(MESES.index(mes_sel) + 1).zfill(2)
    anio_str = str(anio_sel)
    registros = cargar_registros()
    filtrados = [r for r in registros if r["fecha"].startswith(f"{anio_str}-{mes_num}")]

    if not filtrados:
        st.info(f"No hay registros para {mes_sel} {anio_sel}.")
        return

    resumen = {}
    for r in filtrados:
        n = r["nombre"]
        if n not in resumen:
            resumen[n] = {"dias": 0, "trabajado": 0.0, "balance": 0.0}
        resumen[n]["dias"]      += 1
        resumen[n]["trabajado"] += r["horas_trabajadas"]
        resumen[n]["balance"]   += r["diferencia"]

    df_res = pd.DataFrame([{
        "Técnico":  n,
        "Días":     v["dias"],
        "Trabajado": decimal_a_hhmm(v["trabajado"]),
        "Esperado": decimal_a_hhmm(v["dias"] * JORNADA_BASE),
        "Balance":  decimal_a_hhmm(v["balance"]),
    } for n, v in sorted(resumen.items())])
    st.dataframe(df_res, use_container_width=True, hide_index=True)

    m1, m2, m3 = st.columns(3)
    m1.metric("📅 Total jornadas",  sum(v["dias"] for v in resumen.values()))
    m2.metric("⏱️ Total trabajado", decimal_a_hhmm(sum(v["trabajado"] for v in resumen.values())))
    m3.metric("⚖️ Balance global",  decimal_a_hhmm(sum(v["balance"]   for v in resumen.values())))

    st.markdown("---")
    if st.button("📥 Exportar Excel", use_container_width=True):
        wb   = Workbook()
        azul = PatternFill("solid", fgColor="1E3A8A")
        ws1  = wb.active
        ws1.title = "Resumen"
        for c in ws1.append(["Técnico", "Días", "Trabajado", "Esperado", "Balance"]) or ws1[1]:
            c.font      = Font(bold=True, color="FFFFFF")
            c.fill      = azul
            c.alignment = Alignment(horizontal="center")
        for _, row_data in df_res.iterrows():
            ws1.append(list(row_data))

        ws2 = wb.create_sheet("Detalle")
        ws2.append(["Técnico", "Fecha", "Entrada", "Salida", "Inicio Ruta", "Fin Ruta", "Trabajado", "Balance", "Detalle"])
        for c in ws2[1]:
            c.font      = Font(bold=True, color="FFFFFF")
            c.fill      = azul
            c.alignment = Alignment(horizontal="center")
        for r in sorted(filtrados, key=lambda x: (x["nombre"], x["fecha"])):
            ws2.append([
                r["nombre"], r["fecha"], r["hora_entrada"], r["hora_salida"],
                r.get("inicio_ruta", ""), r.get("fin_ruta", ""),
                decimal_a_hhmm(r["horas_trabajadas"]), decimal_a_hhmm(r["diferencia"]), r.get("detalle", ""),
            ])

        for sheet in (ws1, ws2):
            for col in sheet.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                sheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        st.download_button(
            f"⬇️ Descargar Resumen_{mes_num}_{anio_str}.xlsx", buf,
            f"Resumen_{mes_num}_{anio_str}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


# ─────────────────────────────────────────────
# Estadísticas
# ─────────────────────────────────────────────
def pagina_estadisticas():
    st.markdown("### 📈 Estadísticas")
    registros = cargar_registros()
    if not registros:
        st.info("No hay registros para mostrar estadísticas.")
        return

    empleados     = cargar_empleados(solo_activos=False)
    nombres_todos = ["Todos"] + [e["nombre"] for e in empleados]
    col1, col2 = st.columns(2)
    with col1:
        filtro_emp = st.selectbox("Técnico", nombres_todos, key="est_emp")
    with col2:
        anio_sel = st.selectbox("Año", [2024, 2025, 2026, 2027], index=2, key="est_anio")

    filtrados = [r for r in registros if r["fecha"].startswith(str(anio_sel))]
    if filtro_emp != "Todos":
        filtrados = [r for r in filtrados if r["nombre"] == filtro_emp]
    if not filtrados:
        st.info("No hay datos con ese filtro.")
        return

    total_dias    = len(filtrados)
    total_horas   = sum(r["horas_trabajadas"] for r in filtrados)
    total_balance = sum(r["diferencia"]       for r in filtrados)
    promedio      = total_horas / total_dias if total_dias else 0

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("📅 Jornadas",         total_dias)
    m2.metric("⏱️ Total horas",      decimal_a_hhmm(total_horas))
    m3.metric("📊 Promedio diario",   decimal_a_hhmm(promedio))
    m4.metric("⚖️ Balance acumulado", decimal_a_hhmm(total_balance))

    st.markdown("---")
    st.markdown("#### Horas por mes")
    por_mes = {}
    for r in filtrados:
        nm = MESES[int(r["fecha"][5:7]) - 1]
        if nm not in por_mes:
            por_mes[nm] = {"trabajado": 0.0, "balance": 0.0, "dias": 0}
        por_mes[nm]["trabajado"] += r["horas_trabajadas"]
        por_mes[nm]["balance"]   += r["diferencia"]
        por_mes[nm]["dias"]      += 1

    por_mes_ord = {MESES[i]: por_mes[MESES[i]] for i in range(12) if MESES[i] in por_mes}
    df_mes = pd.DataFrame([{
        "Mes": k, "Días": v["dias"],
        "Horas trabajadas": round(v["trabajado"], 2),
        "Balance (h)":      round(v["balance"],   2),
    } for k, v in por_mes_ord.items()])
    st.bar_chart(df_mes.set_index("Mes")[["Horas trabajadas"]])
    st.dataframe(df_mes, use_container_width=True, hide_index=True)

    if filtro_emp == "Todos":
        st.markdown("---")
        st.markdown("#### Ranking por técnico")
        por_tec = {}
        for r in filtrados:
            n = r["nombre"]
            if n not in por_tec:
                por_tec[n] = {"trabajado": 0.0, "dias": 0, "balance": 0.0}
            por_tec[n]["trabajado"] += r["horas_trabajadas"]
            por_tec[n]["dias"]      += 1
            por_tec[n]["balance"]   += r["diferencia"]
        df_tec = pd.DataFrame([{
            "Técnico": n, "Jornadas": v["dias"],
            "Total horas": round(v["trabajado"], 2),
            "Balance (h)": round(v["balance"],   2),
        } for n, v in sorted(por_tec.items(), key=lambda x: -x[1]["trabajado"])])
        st.bar_chart(df_tec.set_index("Técnico")[["Total horas"]])
        st.dataframe(df_tec, use_container_width=True, hide_index=True)


# ─────────────────────────────────────────────
# Técnicos
# ─────────────────────────────────────────────
def pagina_tecnicos():
    st.markdown("### 👷 Técnicos")
    tab1, tab2 = st.tabs(["👥 Ver / Editar", "➕ Nuevo técnico"])

    with tab1:
        empleados = cargar_empleados(solo_activos=False)
        if not empleados:
            st.info("No hay técnicos cargados todavía.")
        else:
            for emp in empleados:
                estado = "🟢 Activo" if emp["activo"] else "🔴 Inactivo"
                with st.expander(f"**{emp['nombre']}** — {estado}", expanded=False):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.markdown(f"📞 **Tel:** {emp['telefono'] or '—'}")
                        st.markdown(f"🪪 **DNI:** {emp['dni'] or '—'}")
                        st.markdown(f"📍 **Zona:** {emp['zona'] or '—'}")
                    with col2:
                        st.markdown(f"🚗 **Vehículo:** {emp['vehiculo'] or '—'}")
                        st.markdown(f"🔖 **Patente:** {emp['patente'] or '—'}")
                    if emp["observaciones"]:
                        st.markdown(f"📝 **Obs:** {emp['observaciones']}")
                    if st.button(f"✏️ Editar {emp['nombre']}", key=f"edit_{emp['id']}"):
                        st.session_state["editar_emp"] = emp.copy()

        if "editar_emp" in st.session_state:
            e = st.session_state["editar_emp"]
            st.markdown(f"---\n**✏️ Editando — {e['nombre']}**")
            with st.form("form_edit_emp"):
                c1, c2 = st.columns(2)
                with c1:
                    e_nombre = st.text_input("Nombre",    value=e["nombre"])
                    e_tel    = st.text_input("Teléfono",  value=e.get("telefono", ""))
                    e_dni    = st.text_input("DNI",       value=e.get("dni", ""))
                    e_zona   = st.text_input("Zona",      value=e.get("zona", ""))
                with c2:
                    e_veh    = st.text_input("Vehículo",  value=e.get("vehiculo", ""))
                    e_pat    = st.text_input("Patente",   value=e.get("patente", ""))
                    e_activo = st.checkbox("Activo",      value=e.get("activo", True))
                e_obs = st.text_area("Observaciones", value=e.get("observaciones", ""))
                if st.form_submit_button("💾 Guardar cambios", use_container_width=True):
                    if not e_nombre:
                        st.error("El nombre no puede estar vacío.")
                    else:
                        actualizar_empleado(e["id"], {
                            "nombre": e_nombre, "activo": e_activo, "telefono": e_tel,
                            "dni": e_dni, "zona": e_zona, "vehiculo": e_veh,
                            "patente": e_pat, "observaciones": e_obs,
                        })
                        del st.session_state["editar_emp"]
                        st.success("Técnico actualizado.")
                        st.rerun()

    with tab2:
        st.markdown("#### Nuevo técnico")
        with st.form("form_nuevo_emp", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                n_nombre = st.text_input("Nombre *",  placeholder="Maxi")
                n_tel    = st.text_input("Teléfono",  placeholder="11 1234-5678")
                n_dni    = st.text_input("DNI",       placeholder="12345678")
                n_zona   = st.text_input("Zona",      placeholder="CABA / Zona Oeste")
            with c2:
                n_veh    = st.text_input("Vehículo",  placeholder="Partner / Moto")
                n_pat    = st.text_input("Patente",   placeholder="AB123CD")
                n_activo = st.checkbox("Activo", value=True)
            n_obs = st.text_area("Observaciones")
            if st.form_submit_button("💾 Guardar técnico", use_container_width=True):
                if not n_nombre:
                    st.error("El nombre es obligatorio.")
                else:
                    guardar_empleado_nuevo({
                        "nombre": n_nombre, "activo": n_activo, "telefono": n_tel,
                        "dni": n_dni, "zona": n_zona, "vehiculo": n_veh,
                        "patente": n_pat, "observaciones": n_obs,
                    })
                    st.success(f"✅ Técnico '{n_nombre}' agregado.")
                    st.rerun()
